import xml.dom.minidom
from openpyxl import Workbook

dom = xml.dom.minidom.parse('go_obo.xml')

root = dom.documentElement

terms = root.getElementsByTagName('term')  # Get all the 'term' elements

workbook = Workbook()  # Create an Excel workbook
sheet = workbook.active


def count_child_nodes(go_id):
    count = 0
    for term in terms:
        id_element = term.getElementsByTagName('id')[0]
        term_id = id_element.firstChild.data
        is_a_elements = term.getElementsByTagName('is_a')

        for is_a_element in is_a_elements:
            is_a_id = is_a_element.firstChild.data
            if is_a_id == go_id:
                count += 1
                count += count_child_nodes(term_id)
    return count


sheet.cell(row=1, column=1, value='GO id')  # Set the column headers
sheet.cell(row=1, column=2, value='Term Name')
sheet.cell(row=1, column=3, value='Definition')
sheet.cell(row=1, column=4, value='Number of Child Nodes')

i = 2

# Iterate over each term element
for index, term in enumerate(terms):
    if "autophagosome" in term.getElementsByTagName('def')[0].getElementsByTagName('defstr')[0].firstChild.data:
        go_id = term.getElementsByTagName('id')[0].firstChild.data
        term_name = term.getElementsByTagName('name')[0].firstChild.data
        definition = term.getElementsByTagName('def')[0].getElementsByTagName('defstr')[0].firstChild.data
        child_nodes = count_child_nodes(go_id)

        # Write the data to the spreadsheet
        sheet.cell(row=i, column=1, value=go_id)
        sheet.cell(row=i, column=2, value=term_name)
        sheet.cell(row=i, column=3, value=definition)
        sheet.cell(row=i, column=4, value=child_nodes)
        i += 1
# Save the workbook
workbook.save('autophagosome.xlsx')
